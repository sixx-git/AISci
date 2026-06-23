"""通用表格文件抽取 — CSV / TSV / XLSX"""
from __future__ import annotations

import csv
import os
from typing import Any, Dict, List

from app.skills.base import BaseSkill, SkillResult
from app.skills.data_finder._utils import new_id


class TabularFileExtractionSkill(BaseSkill):
    name = "TabularFileExtraction"
    description = "从 CSV/TSV/XLSX 导出标准 CSV"

    async def run(self, input_data: Dict[str, Any], context: Dict[str, Any]) -> SkillResult:
        result = SkillResult(success=True)
        file_path = input_data.get("file_path", "")
        source_title = input_data.get("source_title", "")
        output_dir = input_data.get("output_dir", "")

        if not file_path or not os.path.exists(file_path):
            result.add_error("文件不存在")
            result.data = {"tables": []}
            return result

        ext = os.path.splitext(file_path)[1].lower()
        tables: List[Dict[str, Any]] = []

        try:
            if ext in {".csv", ".tsv", ".txt"}:
                tables = [self._from_csv(file_path, source_title, output_dir, delimiter="\t" if ext == ".tsv" else ",")]
            elif ext in {".xlsx", ".xls"}:
                tables = self._from_excel(file_path, source_title, output_dir)
            else:
                result.add_warning(f"不支持的表格格式: {ext}")
        except Exception as exc:
            result.add_error(str(exc))
            result.data = {"tables": [], "errors": [str(exc)]}
            return result

        tables = [t for t in tables if t]
        result.data = {"tables": tables}
        if not tables:
            result.add_warning("未能解析表格文件")
        return result

    def _from_csv(
        self,
        file_path: str,
        source_title: str,
        output_dir: str,
        delimiter: str = ",",
    ) -> Dict[str, Any]:
        with open(file_path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f, delimiter=delimiter)
            columns = list(reader.fieldnames or [])
            rows = list(reader)

        table_id = new_id("tbl")
        out_dir = output_dir or os.path.dirname(file_path)
        os.makedirs(out_dir, exist_ok=True)
        csv_path = os.path.join(out_dir, f"{table_id}.csv")

        if file_path != csv_path:
            with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=columns)
                writer.writeheader()
                writer.writerows(rows)
        else:
            csv_path = file_path

        return {
            "table_id": table_id,
            "source_title": source_title,
            "caption": os.path.basename(file_path),
            "csv_path": csv_path,
            "columns": columns,
            "row_count": len(rows),
            "quality_score": min(1.0, 0.5 + 0.1 * len(columns)),
            "extraction_method": "tabular_file",
            "source_type": "tabular_file",
        }

    def _from_excel(
        self,
        file_path: str,
        source_title: str,
        output_dir: str,
    ) -> List[Dict[str, Any]]:
        try:
            import openpyxl
        except ImportError:
            raise ValueError("需要 openpyxl 解析 xlsx") from None

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        out_dir = output_dir or os.path.dirname(file_path)
        os.makedirs(out_dir, exist_ok=True)
        tables: List[Dict[str, Any]] = []

        for sheet_name in wb.sheetnames[:3]:
            ws = wb[sheet_name]
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header = next(rows_iter)
            except StopIteration:
                continue
            columns = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(header)]
            if not any(columns):
                continue
            data_rows = []
            for row in rows_iter:
                if row is None:
                    continue
                data_rows.append({columns[i]: ("" if v is None else str(v)) for i, v in enumerate(row) if i < len(columns)})
                if len(data_rows) >= 5000:
                    break

            table_id = new_id("tbl")
            csv_path = os.path.join(out_dir, f"{table_id}.csv")
            with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=columns)
                writer.writeheader()
                writer.writerows(data_rows)

            tables.append({
                "table_id": table_id,
                "source_title": source_title,
                "caption": f"{os.path.basename(file_path)}::{sheet_name}",
                "csv_path": csv_path,
                "columns": columns,
                "row_count": len(data_rows),
                "quality_score": 0.75,
                "extraction_method": "excel_sheet",
                "source_type": "tabular_file",
            })
        wb.close()
        return tables
